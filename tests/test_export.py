import openpyxl
from io import BytesIO

import export


def _sample_scores():
    empty = {"score": None, "breakdown": []}
    return {"demand": empty, "financial": empty, "marketing": empty, "risk": empty, "overall": empty}


def _sample_args():
    return dict(
        talent={"name": "Test Artist", "domain": "music"},
        performance={
            "venue_name": "Test Hall", "city": "Austin", "estimated_date": "2026-01-01",
            "target_capacity": 1000, "budget": 50000.0, "notes": "some notes",
        },
        revenue_info={
            "ticket_price": 45.0, "ticket_price_source": "default",
            "sell_through_rate": 0.75, "sell_through_rate_source": "default",
            "estimated_attendance": 750, "estimated_revenue": 33750.0,
        },
        expense_info={
            "total_expenses": 40000.0,
            "breakdown": {"venue": 10000.0, "marketing": 6000.0, "production": 8000.0,
                          "talent_fee": 12000.0, "other": 4000.0},
        },
        net_margin=-6250.0,
        venue_fit_score=0.75,
        marketing_efficiency=8.0,
        demand={"search_interest_index": 42.0},
        audience={"monthly_listeners": 850000.0},
        financial_details={"artist_guarantee": 20000.0},
        touring_history={"sold_out_similar_venues": True},
        market_competition={"other_concerts_count": 1},
        scores=_sample_scores(),
        similar=[{"comparable_name": "Other Artist", "avg_capacity": 900}],
        historical_summary={"in_city": [{"city": "Austin"}], "elsewhere": []},
    )


def test_build_booking_workbook_produces_all_expected_sheets():
    xlsx_bytes = export.build_booking_workbook(**_sample_args())
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    assert wb.sheetnames == [
        "Booking summary", "Confidence scores", "Score breakdown", "Expense breakdown",
        "Demand metrics", "Audience & social media", "Financial details",
        "Touring history", "Market competition", "Similar talent",
        "History - in city", "History - elsewhere",
    ]


def test_build_booking_workbook_summary_sheet_has_key_fields():
    xlsx_bytes = export.build_booking_workbook(**_sample_args())
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Booking summary"]
    fields = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "Talent name" in fields
    assert "Estimated revenue ($)" in fields
    assert "Venue fit score" in fields


def test_build_booking_workbook_confidence_scores_sheet():
    args = _sample_args()
    args["scores"] = {
        "demand": {"score": 87, "breakdown": [{"label": "Monthly listeners", "raw_value": 850000,
                                                "sub_score": 80, "note": "n/a"}]},
        "financial": {"score": 74, "breakdown": []},
        "marketing": {"score": 91, "breakdown": []},
        "risk": {"score": 32, "breakdown": []},
        "overall": {"score": 83, "breakdown": []},
    }
    xlsx_bytes = export.build_booking_workbook(**args)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    scores_ws = wb["Confidence scores"]
    values = {row[0]: row[1] for row in scores_ws.iter_rows(min_row=2, values_only=True)}
    assert values["Demand"] == 87
    assert values["Risk (lower is better)"] == 32
    breakdown_ws = wb["Score breakdown"]
    breakdown_rows = list(breakdown_ws.iter_rows(min_row=2, values_only=True))
    assert any(row[1] == "Monthly listeners" for row in breakdown_rows)


def test_build_booking_workbook_handles_all_empty_sections():
    args = _sample_args()
    args["demand"] = {}
    args["audience"] = None
    args["financial_details"] = None
    args["touring_history"] = None
    args["market_competition"] = None
    args["similar"] = []
    xlsx_bytes = export.build_booking_workbook(**args)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    assert wb["Demand metrics"].cell(row=2, column=1).value == "(none entered)"
    assert wb["Audience & social media"].cell(row=2, column=1).value == "(none entered)"
    assert wb["Financial details"].cell(row=2, column=1).value == "(none entered)"
    assert wb["Touring history"].cell(row=2, column=1).value == "(none entered)"
    assert wb["Market competition"].cell(row=2, column=1).value == "(none entered)"
